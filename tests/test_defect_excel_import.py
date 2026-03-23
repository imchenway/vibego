from __future__ import annotations

import asyncio
import os
import sys
from datetime import datetime
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.base import StorageKey
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import BufferedInputFile, InlineKeyboardMarkup
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

os.environ.setdefault("BOT_TOKEN", "TEST_TOKEN")

import bot  # noqa: E402

LATEST_TEMPLATE_HEADERS = [
    "缺陷标题",
    "前置条件",
    "复现步骤",
    "现状",
    "预期效果",
    "关联任务编码",
    "优先级",
    "附件",
]

PREVIOUS_TEMPLATE_HEADERS = [
    "缺陷标题",
    "前置条件",
    "复现步骤",
    "现状",
    "预期效果",
    "关联任务编码",
    "优先级",
]

LEGACY_TEMPLATE_HEADERS = [
    "缺陷标题",
    "前置条件",
    "复现步骤",
    "预期结果",
    "关联任务编码",
    "优先级",
]


class DummyMessage:
    def __init__(self, *, chat_id: int = 1, user_id: int = 1, text: str | None = None):
        self.calls = []
        self.edits = []
        self.documents = []
        self.chat = SimpleNamespace(id=chat_id)
        self.from_user = SimpleNamespace(id=user_id, full_name="Tester")
        self.message_id = 100
        self.date = datetime.now(bot.UTC)
        self.text = text
        self.caption = None
        self.document = None

    async def answer(self, text: str, parse_mode=None, reply_markup=None, **kwargs):
        self.calls.append((text, parse_mode, reply_markup, kwargs))
        return SimpleNamespace(message_id=self.message_id + len(self.calls), chat=self.chat)

    async def edit_text(self, text: str, parse_mode=None, reply_markup=None, **kwargs):
        self.edits.append((text, parse_mode, reply_markup, kwargs))
        return SimpleNamespace(message_id=self.message_id, chat=self.chat)

    async def answer_document(self, document, caption=None, **kwargs):
        self.documents.append((document, caption, kwargs))
        return SimpleNamespace(message_id=self.message_id + len(self.documents), chat=self.chat)


class DummyCallback:
    def __init__(self, data: str, message: DummyMessage):
        self.data = data
        self.message = message
        self.answers = []
        self.from_user = SimpleNamespace(id=1, full_name="Tester")

    async def answer(self, text: str | None = None, show_alert: bool = False):
        self.answers.append((text, show_alert))


def make_state(message: DummyMessage) -> tuple[FSMContext, MemoryStorage]:
    storage = MemoryStorage()
    state = FSMContext(
        storage=storage,
        key=StorageKey(bot_id=999, chat_id=message.chat.id, user_id=message.from_user.id),
    )
    return state, storage


def _build_workbook(path: Path, rows: list[list[object]]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "缺陷导入模板"
    ws.append(LATEST_TEMPLATE_HEADERS)
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


def test_task_list_view_includes_excel_import_button(monkeypatch):
    class DummyService:
        async def paginate(self, **kwargs):
            return [], 1

        async def count_tasks(self, **kwargs):
            return 0

    monkeypatch.setattr(bot, "TASK_SERVICE", DummyService())

    text, markup = asyncio.run(bot._build_task_list_view(status=None, page=1, limit=10))

    assert text.startswith("*任务列表*")
    buttons = [button.text for row in markup.inline_keyboard for button in row]
    assert "📥 Excel批量创建缺陷" in buttons


def test_task_defect_excel_import_callback_enters_import_view():
    message = DummyMessage()
    callback = DummyCallback(bot.TASK_DEFECT_EXCEL_IMPORT_CALLBACK, message)
    bot._init_task_view_context(message, bot._make_list_view_state(status=None, page=1, limit=10))

    async def _scenario() -> None:
        await bot.on_task_defect_excel_import(callback)
        state = bot._peek_task_view(message.chat.id, message.message_id)
        assert state is not None
        assert state.kind == "defect_excel_import"
        assert message.edits
        text, _parse_mode, markup, _kwargs = message.edits[-1]
        assert "Excel 批量创建缺陷" in text
        assert isinstance(markup, InlineKeyboardMarkup)

    asyncio.run(_scenario())


def test_task_defect_excel_template_callback_sends_xlsx_document():
    message = DummyMessage()
    callback = DummyCallback(bot.TASK_DEFECT_EXCEL_TEMPLATE_CALLBACK, message)

    async def _scenario() -> None:
        await bot.on_task_defect_excel_template(callback)
        assert callback.answers[-1] == ("模板已发送", False)
        assert message.documents
        document, caption, _kwargs = message.documents[-1]
        assert isinstance(document, BufferedInputFile)
        assert document.filename == bot.DEFECT_EXCEL_TEMPLATE_FILENAME
        workbook = load_workbook(filename=BytesIO(document.data))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        assert headers[: len(LATEST_TEMPLATE_HEADERS)] == LATEST_TEMPLATE_HEADERS
        assert "请按模板填写后再上传" in (caption or "")

    asyncio.run(_scenario())


def test_task_defect_excel_upload_valid_file_enters_confirm(monkeypatch, tmp_path: Path):
    workbook_path = _build_workbook(
        tmp_path / "defects.xlsx",
        [[
            "登录按钮无响应",
            "已登录测试账号",
            "1. 打开页面",
            "页面无跳转",
            "页面应进入首页",
            "",
            3,
            "https://cdn.example.com/errors/login.png\nartifacts/error.log\nhttps://cdn.example.com/errors/login.png",
        ]],
    )
    message = DummyMessage()
    message.document = SimpleNamespace(file_name="defects.xlsx")
    state, _storage = make_state(message)
    asyncio.run(
        state.update_data(
            origin_message=message,
            origin_status=None,
            origin_page=1,
            origin_limit=10,
            actor="Tester#1",
        )
    )
    asyncio.run(state.set_state(bot.TaskDefectExcelImportStates.waiting_upload))

    async def fake_collect(_message, _target_dir):
        return [
            bot.TelegramSavedAttachment(
                kind="document",
                display_name="defects.xlsx",
                mime_type=bot.DEFECT_EXCEL_MIME,
                absolute_path=workbook_path,
                relative_path="./data/defects.xlsx",
            )
        ]

    monkeypatch.setattr(bot, "_collect_saved_attachments", fake_collect)

    async def _scenario() -> None:
        await bot.on_task_defect_excel_upload(message, state)
        assert await state.get_state() == bot.TaskDefectExcelImportStates.waiting_confirm.state
        data = await state.get_data()
        assert len(data["validated_rows"]) == 1
        assert data["validated_rows"][0]["precondition"] == "已登录测试账号"
        assert data["validated_rows"][0]["current_state"] == "页面无跳转"
        assert data["validated_rows"][0]["pending_attachments"] == [
            {
                "kind": "document",
                "display_name": "login.png",
                "mime_type": "image/png",
                "path": "https://cdn.example.com/errors/login.png",
            },
            {
                "kind": "document",
                "display_name": "error.log",
                "mime_type": "text/plain",
                "path": "artifacts/error.log",
            },
        ]
        assert message.calls
        assert "Excel 预检通过" in message.calls[-1][0]

    asyncio.run(_scenario())


def test_task_defect_excel_upload_previous_headers_without_attachment_still_enter_confirm(monkeypatch, tmp_path: Path):
    workbook_path = tmp_path / "previous_defects.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "缺陷导入模板"
    ws.append(PREVIOUS_TEMPLATE_HEADERS)
    ws.append(["登录按钮无响应", "已登录测试账号", "1. 打开页面", "页面无跳转", "页面应进入首页", "", 3])
    wb.save(workbook_path)

    message = DummyMessage()
    message.document = SimpleNamespace(file_name="previous_defects.xlsx")
    state, _storage = make_state(message)
    asyncio.run(
        state.update_data(
            origin_message=message,
            origin_status=None,
            origin_page=1,
            origin_limit=10,
            actor="Tester#1",
        )
    )
    asyncio.run(state.set_state(bot.TaskDefectExcelImportStates.waiting_upload))

    async def fake_collect(_message, _target_dir):
        return [
            bot.TelegramSavedAttachment(
                kind="document",
                display_name="previous_defects.xlsx",
                mime_type=bot.DEFECT_EXCEL_MIME,
                absolute_path=workbook_path,
                relative_path="./data/previous_defects.xlsx",
            )
        ]

    monkeypatch.setattr(bot, "_collect_saved_attachments", fake_collect)

    async def _scenario() -> None:
        await bot.on_task_defect_excel_upload(message, state)
        assert await state.get_state() == bot.TaskDefectExcelImportStates.waiting_confirm.state
        data = await state.get_data()
        assert len(data["validated_rows"]) == 1
        assert data["validated_rows"][0]["current_state"] == "页面无跳转"
        assert data["validated_rows"][0]["expected_effect"] == "页面应进入首页"
        assert data["validated_rows"][0]["pending_attachments"] == []

    asyncio.run(_scenario())


def test_task_defect_excel_upload_legacy_headers_still_enter_confirm(monkeypatch, tmp_path: Path):
    workbook_path = tmp_path / "legacy_defects.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "缺陷导入模板"
    ws.append(LEGACY_TEMPLATE_HEADERS)
    ws.append(["登录按钮无响应", "已登录测试账号", "1. 打开页面", "页面应进入首页", "", 3])
    wb.save(workbook_path)

    message = DummyMessage()
    message.document = SimpleNamespace(file_name="legacy_defects.xlsx")
    state, _storage = make_state(message)
    asyncio.run(
        state.update_data(
            origin_message=message,
            origin_status=None,
            origin_page=1,
            origin_limit=10,
            actor="Tester#1",
        )
    )
    asyncio.run(state.set_state(bot.TaskDefectExcelImportStates.waiting_upload))

    async def fake_collect(_message, _target_dir):
        return [
            bot.TelegramSavedAttachment(
                kind="document",
                display_name="legacy_defects.xlsx",
                mime_type=bot.DEFECT_EXCEL_MIME,
                absolute_path=workbook_path,
                relative_path="./data/legacy_defects.xlsx",
            )
        ]

    monkeypatch.setattr(bot, "_collect_saved_attachments", fake_collect)

    async def _scenario() -> None:
        await bot.on_task_defect_excel_upload(message, state)
        assert await state.get_state() == bot.TaskDefectExcelImportStates.waiting_confirm.state
        data = await state.get_data()
        assert len(data["validated_rows"]) == 1
        assert data["validated_rows"][0]["current_state"] == ""
        assert data["validated_rows"][0]["expected_effect"] == "页面应进入首页"
        assert data["validated_rows"][0]["pending_attachments"] == []

    asyncio.run(_scenario())


def test_task_defect_excel_upload_invalid_file_keeps_upload_state(monkeypatch, tmp_path: Path):
    workbook_path = tmp_path / "invalid.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["错误表头"])
    wb.save(workbook_path)

    message = DummyMessage()
    message.document = SimpleNamespace(file_name="invalid.xlsx")
    state, _storage = make_state(message)
    asyncio.run(
        state.update_data(
            origin_message=message,
            origin_status=None,
            origin_page=1,
            origin_limit=10,
            actor="Tester#1",
        )
    )
    asyncio.run(state.set_state(bot.TaskDefectExcelImportStates.waiting_upload))

    async def fake_collect(_message, _target_dir):
        return [
            bot.TelegramSavedAttachment(
                kind="document",
                display_name="invalid.xlsx",
                mime_type=bot.DEFECT_EXCEL_MIME,
                absolute_path=workbook_path,
                relative_path="./data/invalid.xlsx",
            )
        ]

    monkeypatch.setattr(bot, "_collect_saved_attachments", fake_collect)

    async def _scenario() -> None:
        await bot.on_task_defect_excel_upload(message, state)
        assert await state.get_state() == bot.TaskDefectExcelImportStates.waiting_upload.state
        assert message.calls
        assert "Excel 预检失败" in message.calls[-1][0]

    asyncio.run(_scenario())


def test_task_defect_excel_upload_invalid_attachment_ref_keeps_upload_state(monkeypatch, tmp_path: Path):
    workbook_path = _build_workbook(
        tmp_path / "invalid_attachment.xlsx",
        [[
            "登录按钮无响应",
            "已登录测试账号",
            "1. 打开页面",
            "页面无跳转",
            "页面应进入首页",
            "",
            3,
            "ftp://example.com/login.png",
        ]],
    )

    message = DummyMessage()
    message.document = SimpleNamespace(file_name="invalid_attachment.xlsx")
    state, _storage = make_state(message)
    asyncio.run(
        state.update_data(
            origin_message=message,
            origin_status=None,
            origin_page=1,
            origin_limit=10,
            actor="Tester#1",
        )
    )
    asyncio.run(state.set_state(bot.TaskDefectExcelImportStates.waiting_upload))

    async def fake_collect(_message, _target_dir):
        return [
            bot.TelegramSavedAttachment(
                kind="document",
                display_name="invalid_attachment.xlsx",
                mime_type=bot.DEFECT_EXCEL_MIME,
                absolute_path=workbook_path,
                relative_path="./data/invalid_attachment.xlsx",
            )
        ]

    monkeypatch.setattr(bot, "_collect_saved_attachments", fake_collect)

    async def _scenario() -> None:
        await bot.on_task_defect_excel_upload(message, state)
        assert await state.get_state() == bot.TaskDefectExcelImportStates.waiting_upload.state
        assert message.calls
        assert "第 2 行：附件引用不合法" in message.calls[-1][0]

    asyncio.run(_scenario())


def test_task_defect_excel_confirm_creates_defects_and_restores_list(monkeypatch):
    message = DummyMessage(text="✅ 确认创建")
    origin_message = DummyMessage()
    state, _storage = make_state(message)
    asyncio.run(
        state.update_data(
            origin_message=origin_message,
            origin_status=None,
            origin_page=1,
            origin_limit=10,
            actor="Tester#1",
            imported_file_name="defects.xlsx",
            imported_total_rows=2,
            validated_rows=[
                {
                    "title": "缺陷一",
                    "precondition": "已登录测试账号",
                    "reproduction": "步骤1",
                    "current_state": "页面无反馈",
                    "expected_effect": "结果1",
                    "related_task_id": None,
                    "priority": 3,
                    "pending_attachments": [
                        {
                            "kind": "document",
                            "display_name": "login.png",
                            "mime_type": "image/png",
                            "path": "https://cdn.example.com/errors/login.png",
                        },
                        {
                            "kind": "document",
                            "display_name": "error.log",
                            "mime_type": "text/plain",
                            "path": "artifacts/error.log",
                        },
                    ],
                },
                {
                    "title": "缺陷二",
                    "precondition": "",
                    "reproduction": "",
                    "current_state": "",
                    "expected_effect": "",
                    "related_task_id": None,
                    "priority": 2,
                    "pending_attachments": [],
                },
            ],
        )
    )
    asyncio.run(state.set_state(bot.TaskDefectExcelImportStates.waiting_confirm))

    created: list[dict[str, object]] = []

    async def fake_create_root_task(*, title, status, priority, task_type, tags, due_date, description, related_task_id, actor):
        created.append({"title": title, "priority": priority, "description": description})
        return SimpleNamespace(id=f"TASK_{len(created):04d}")

    async def fake_build_task_list_view(*, status, page, limit):
        return "*任务列表*", InlineKeyboardMarkup(inline_keyboard=[])

    bound: list[tuple[str, list[dict[str, str]], str]] = []

    async def fake_bind_serialized_attachments(task, attachments, *, actor):
        bound.append((task.id, list(attachments), actor))
        return []

    monkeypatch.setattr(bot.TASK_SERVICE, "create_root_task", fake_create_root_task)
    monkeypatch.setattr(bot, "_build_task_list_view", fake_build_task_list_view)
    monkeypatch.setattr(bot, "_bind_serialized_attachments", fake_bind_serialized_attachments)

    async def _scenario() -> None:
        await bot.on_task_defect_excel_confirm(message, state)
        assert await state.get_state() is None
        assert created[0]["title"] == "缺陷一"
        assert created[0]["priority"] == 3
        assert (
            created[0]["description"]
            == "前置条件：\n已登录测试账号\n\n复现步骤：\n步骤1\n\n现状：\n页面无反馈\n\n预期效果：\n结果1"
        )
        assert created[1]["title"] == "缺陷二"
        assert created[1]["priority"] == 2
        assert created[1]["description"] == "前置条件：\n-\n\n复现步骤：\n-\n\n现状：\n-\n\n预期效果：\n-"
        assert bound == [
            (
                "TASK_0001",
                [
                    {
                        "kind": "document",
                        "display_name": "login.png",
                        "mime_type": "image/png",
                        "path": "https://cdn.example.com/errors/login.png",
                    },
                    {
                        "kind": "document",
                        "display_name": "error.log",
                        "mime_type": "text/plain",
                        "path": "artifacts/error.log",
                    },
                ],
                "Tester#1",
            )
        ]
        assert origin_message.edits, "创建完成后应恢复原任务列表"
        assert message.calls
        assert "Excel 导入结果" in message.calls[-1][0]

    # DummyMessage in this file has no edit fallback helper context, so provide edit ability through origin_message
    origin_message.edits = []
    asyncio.run(_scenario())
